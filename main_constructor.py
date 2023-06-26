import configparser
import csv
import os
import re
import shutil
import smtplib
from datetime import datetime as dt
from dataclasses import dataclass
from string import Template

import openpyxl as openpyxl

from send_mail import SendMail


class DataInFileError(Exception):
    def __init__(self, *args):
        if args:
            self.message = args[0]
        else:
            self.message = None

    def __str__(self):
        if self.message:
            return 'DataInFileError, {0} '.format(self.message)
        else:
            return 'DataInFileError has been raised'


class FileHTMLError(Exception):
    def __init__(self, *args):
        if args:
            self.message = args[0]
        else:
            self.message = None

    def __str__(self):
        if self.message:
            return 'FileXlsError, {0} '.format(self.message)
        else:
            return 'FileXlsError has been raised'


class FileXlsError(Exception):
    def __init__(self, *args):
        if args:
            self.message = args[0]
        else:
            self.message = None

    def __str__(self):
        if self.message:
            return 'FileXlsError, {0} '.format(self.message)
        else:
            return 'FileXlsError has been raised'


@dataclass()
class LogMessage:
    def message_log(self, unique, mes, type_mes='INFO'):
        now = dt.now()
        time_log = now.strftime("%d.%m.%Y-%H_%M_%S")
        time_for_file = now.strftime("%d.%m.%Y")
        mes_data = [
            [unique, mes, type_mes, time_log]
        ]
        log_dir = os.path.join(os.getcwd(), 'log')
        if not os.path.exists(log_dir):
            os.mkdir(log_dir)
        with open('log\log_send_mail_' + time_for_file + '.csv', 'a', encoding='cp1251',
                  newline="") as file:
            # !!!! чтобы в windows разделитель в виде ; bcgjkmpet
            writer = csv.writer(file, delimiter=";")
            writer.writerows(
                mes_data
            )


@dataclass()
class CreateMail(LogMessage):
    """
    Создание письма на основе ????
    file_exclude_email: Файл в котором формируется список для исключения двойной отправки
    """
    file_patten: str
    files_in_dir: str
    file_exclude_email: str
    head_row_in_files: int
    column_email_in_files: list
    pattern_for_email_re: str = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b'
    theme_mail: str = None
    id_image_include: str = None
    file_image_for_include: list = None
    id_image_for_include: str = None
    files_image_for_include: list = None

    def __post_init__(self):
        # Создаем множество(массив уникальных не индексируемых элементов)
        self.set_exclude_email = set()
        self.set_email_to = set()
        absFilePath = os.path.abspath(__file__)
        path, filename = os.path.split(absFilePath)
        print('Директория содержащая модуль:', path)
        self.path_module = path
        # message = LogMessage(path)
        # self.message_log = message.message_log

    # Читаем файл шаблона и входной файл для отправки.
    # Если идет расхождение в заголовке и в шаблонном
    # файле прерываем обработку до устранения расхождений

    def main(self):
        try:
            # Читаем файл шаблона и заносим в словарь.
            validation_html = self.read_and_check_pattern_html(os.path.join(self.path_module, self.file_patten))
            if not validation_html[0]:
                raise FileHTMLError(
                    'Ошибка при Валидации HTML шаблона'
                )
            self.read_in_files(validation_html[1])
            self.message_log(
                '',
                'Обработка завершенна'
            )
        except(Exception,) as er:
            self.message_log(
                '',
                er,
                'error'
            )

    def read_in_files(self, pattern: str):
        """
        Читаем и проверяем входящие файлы на соответствие структуре данных
        :param pattern:
        :return:
        """
        files_array = self.array_files(self.files_in_dir)
        # print(self.sender_email)
        # print(type(self.sender_email))
        # Просматриваем файл исключений(файл для записи емейлов с ранее уже отправленными на эти адреса письмами)
        file_path_exclude = os.path.join(self.path_module, self.file_exclude_email)
        try:
            with open(file_path_exclude) as file:
                list_emails = file.read().splitlines()
                self.set_exclude_email.update(list_emails)
        except(FileNotFoundError,):
            print('Не найден файл исключения. Будут отработанны все e-mail')
            pass
        for i_file in files_array:
            print(i_file)
            try:
                wb = openpyxl.load_workbook(i_file)
            except(Exception,) as er:
                self.message_log(
                    '',
                    f'Не удалось открыть файл:{i_file}.'
                    ' Файл не будет обработан.', 'error'
                )
                continue
            sheet_with_data = wb.worksheets[0]
            # Начинаем обрабатывать записи в цикле
            # Проверяем на наличие записей:
            count_row = sheet_with_data.max_row + 1
            print(count_row)
            print(self.head_row_in_files)
            point_next_note = False
            for item_row in range(self.head_row_in_files + 1, count_row):
                # # Изымаем и проверяем данные построчно
                # Так как в файле может содержаться одновременно в нескольких
                # столбцах e-mail(задаем в settings.ini) и поочереди перебираем
                # его на соответствие:
                value_email = ''
                print('Строка', item_row)
                for i_column in self.column_email_in_files:
                    value_email = str(
                        sheet_with_data.cell(
                            row=item_row, column=int(i_column)).value
                    ).strip().lower()
                    # Чтобы распердлить сообщения об отказе, отсекаем пустые ячейки
                    if value_email == "":
                        continue
                    print('Значение проверяемого e-mail:', value_email)
                    # Проверяем e-mail
                    match_email = re.fullmatch(self.pattern_for_email_re, value_email)
                    if not match_email:
                        print('Некорректный e-mail', value_email)
                        self.message_log(
                            '',
                            f' ячейка со значением {value_email} в  строке {item_row}'
                            f' и столбце:{i_column}, файла {i_file}'
                            f' имеет неверный формат e-mail.')
                        value_email = None
                        continue
                    else:
                        # Сначала сделаем множество(для удаления дубликатов)
                        # а в дальнейшем переведем в список для пропорциональной отправки
                        break
                if value_email in ['', None]:
                    self.message_log(
                        '',
                        f' В строке {item_row}, файла {i_file} не найденны данные об email.')
                    continue
                # Проверяем, а не был ли отправлен ранее на этот емайл сообщение.
                if value_email in self.set_exclude_email:
                    print(f'{value_email} содержится в списке исключений для отправки e-mail')
                    self.message_log(
                        value_email,
                        'Адрес почты содержится в списке исключений для отправки. Отправка не будет произведена.'
                    )
                    continue
                self.set_email_to.add(value_email)
        print(self.set_email_to)
        # Создаем класс для отправки письма и передаем параметры в него
        push_email = SendMail(
            theme_global=self.theme_mail,
            html_global=pattern,
            id_image_for_include=self.id_image_for_include,
            files_image_for_include=list(
                map(lambda x: os.path.join(self.path_module, x),
                    self.files_image_for_include
                    )
            )
        )
        for email in self.set_email_to:
            try:
                # Так как метод для массовой рассылке то подразумевается что приходит список
                push_email.push_mail_group(
                    emails_for_send=[email]
                )
            except(Exception,) as er:
                self.message_log(email, f'В процессе отправки произошла ошибка{er}.', 'ERROR')
                continue
            with open(file_path_exclude, 'a', encoding='utf-8', newline="") as file_csv:

                writer = csv.writer(file_csv, delimiter=";")
                writer.writerow(
                    [email]
                )
            self.message_log(email, 'Успешное завершение отправки сообщения.')
        print('Работа по отправки сообщений завершенна.')

        # Закончил здесь надо формировать текст письма
        # Иногда почтовые клиенты не отображают
        # текст(поэтому делаем альтернативный варинант с текстом)

        # push_email.push_mail(
        #     value_email,
        #     body_for_mail_html,
        #     body_for_mail_text
        # )
        # self.message_log(
        #     '',
        #     f'Работа с записью для {value_email} '
        #     f'в файле: {i_file} успешно завершена.'
        # )
        # self.message_log(
        #     '',
        #     f'Завершена работа с файлом: {i_file}.'
        # )
        # self.file_transfer(i_file, 'success')

    def file_transfer(self, file, directory='error'):
        """
        Переносим файл в папку ошибок
        """
        if not os.path.exists(file):
            raise FileNotFoundError
        error_dir = os.path.join(os.getcwd(), directory)
        if not os.path.exists(error_dir):
            os.mkdir(error_dir)
        file_name = file.split('\\')
        hand_path = os.path.join(error_dir, file_name[-1])
        file_package_path_tmp = shutil.move(file, hand_path)
        self.message_log('',
                         f'Файл: {file} перемещен в '
                         f'каталог {directory} файлов: {file_package_path_tmp}')

    def array_files(self, path_in: str):
        """Формирование списка файлов."""

        def validation_name_file(file):
            if file.endswith('.xlsx') or file.endswith('.xlxm'):
                return os.path.join(dir_path, file_name)
            elif file_name.endswith('.xls'):
                raise FileXlsError(
                    f'Найден файл Excel старого формата(.xls) в каталоге {path_in}.'
                    f' Файлы данного типа не обрабатываются, но могут содержать'
                    f' информацию для отправки. Обработка и отправка остановленна.')

        list_file = []
        for dir_path, dir_names, file_names in os.walk(path_in):
            for file_name in file_names:
                list_file.append(validation_name_file(file_name))
        count_array = len(list_file)
        if count_array == 0:
            raise FileNotFoundError(
                f'В директории "{path_in}" нет файлов для обработки.'
            )
        return list_file


    def read_and_check_pattern_html(self, file_pattern):
        print('Читаем файл html', file_pattern)
        # Проверяем что в передоваемом файле действительно содержится HTML
        from bs4 import BeautifulSoup
        try:
            with open(file_pattern, 'r', encoding='utf-8') as file:
                s = file.read()

        except FileNotFoundError:
            print("Невозможно открыть файл")

        if bool(BeautifulSoup(s, "html.parser").find()):
            # Если прошел валидацию то возвращаем шаблон
            return [True, s]
        else:
            return [False, ]


def start():
    # Читаем параметры с settings.ini
    config = configparser.ConfigParser()
    file_settings = os.path.join(os.getcwd(), 'settings.ini')
    config.read(file_settings, encoding='utf-8-sig')
    # print(config)
    directory_with_xlsx_files = str(
        config['Excel']['directory_with_xlsx_files']
    )
    file_pattens_html = str(
        config['Main']['file_pattens_html']
    )
    file_exclude_email = str(
        config['Main']['exclude_email']
    )
    head_row_in_files = int(
        config['Excel']['head_row_in_files']
    )
    theme_mail = str(
        config['Main']['theme_mail']
    )
    column_email_in_files = config['Excel']['column_email_in_files'].split(',')
    id_image_for_include = str(
        config['Main']['id_image_for_include']
    )
    files_image_for_include = config['Main']['files_image_for_include'].split(',')
    read = CreateMail(
        file_patten=file_pattens_html,
        files_in_dir=directory_with_xlsx_files,
        file_exclude_email=file_exclude_email,
        head_row_in_files=head_row_in_files,
        column_email_in_files=column_email_in_files,
        theme_mail=theme_mail,
        id_image_for_include=id_image_for_include,
        files_image_for_include=files_image_for_include
    )
    read.main()


if __name__ == '__main__':
    start()
